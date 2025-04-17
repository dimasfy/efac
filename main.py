import cv2
import pytesseract
import os
import serial
import time
from openpyxl import load_workbook

# Configure Tesseract OCR
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Update path

# Frame width and height
frameWidth = 1000
frameHeight = 480

# Load Haar Cascade for license plate detection
plateCascade = cv2.CascadeClassifier("model.xml")
minArea = 500

# Initialize video capture
cap = cv2.VideoCapture(0)
cap.set(3, frameWidth)
cap.set(4, frameHeight)
cap.set(10, 150)
count = 0

# Load dataset
def load_dataset(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    return set(str(cell.value).strip() for row in ws.iter_rows() for cell in row if cell.value is not None)

allowed_plates = load_dataset('daftarplat.xlsx')

def read_plate_from_image(image_path):
    img = cv2.imread(image_path)
    img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    plate_thresh = cv2.threshold(img_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    text = pytesseract.image_to_string(plate_thresh, config='--psm 8')
    return text.strip()

ser = serial.Serial('COM36', 9600, timeout=1)  
time.sleep(2)  

while True:
    success, img = cap.read()
    if not success:
        break

    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    numberPlates = plateCascade.detectMultiScale(imgGray, 1.1, 4)
    plate_detected = False
    imgRoi = None

    for (x, y, w, h) in numberPlates:
        area = w * h
        if area > minArea:
            cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
            cv2.putText(img, "NumberPlate", (x, y - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)
            imgRoi = img[y:y + h, x:x + w]
            plate_detected = True
            cv2.imshow("Number Plate", imgRoi)

    cv2.imshow("Result", img)

    # Read distance from Arduino
    if ser.in_waiting > 0:
        distance = ser.readline().decode('utf-8').strip()
        try:
            distance = float(distance)
            if distance < 5 and plate_detected:  
                img_path = f"./IMAGES/{str(count)}.jpg"
                cv2.imwrite(img_path, imgRoi)
                cv2.rectangle(img, (0, 200), (640, 300), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, "Scan Saved", (15, 265), cv2.FONT_HERSHEY_COMPLEX, 2, (0, 0, 255), 2)
                cv2.imshow("Result", img)
                cv2.waitKey(500)

                plate_text = read_plate_from_image(img_path)
                print(f"Detected number plate text: {plate_text}")

                if plate_text in allowed_plates:
                    print("Access Granted: Silahkan Masuk")
                    cv2.putText(img, "Silahkan Masuk", (50, 50), cv2.FONT_HERSHEY_COMPLEX, 2, (0, 255, 0), 2)
                    ser.write(b'O')  
                    print("Portal terbuka")
                else:
                    print("Access Denied")

                count += 1
        except ValueError:
            pass  

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
